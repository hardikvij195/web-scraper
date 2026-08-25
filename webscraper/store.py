"""SQLite persistence: jobs + places. Stdlib sqlite3, one file under data/."""
from __future__ import annotations

import csv
import json
import logging
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from webscraper.config import settings
from webscraper.models import Place

log = logging.getLogger("webscraper.store")

PLACE_COLS = [
    "job_id", "place_key", "name", "category", "address", "country", "phone", "phone_digits", "website", "domain",
    "rating", "reviews_count", "price_range", "lat", "lng", "distance_km", "maps_url", "plus_code", "place_id",
    "email", "emails", "instagram", "facebook", "linkedin", "twitter_x", "youtube", "tiktok",
    "whatsapp_number", "whatsapp_source", "enrich_status", "enriched_at", "scraped_at",
    "summary", "owner", "team", "research_status", "researched_at", "raw",
]
_JSON_COLS = {"emails", "raw", "team"}

SCHEMA = """
CREATE TABLE IF NOT EXISTS jobs (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  query TEXT NOT NULL,
  location TEXT,
  max_places INTEGER,
  delay_sec REAL,
  status TEXT NOT NULL DEFAULT 'running',   -- running | done | stopped | failed
  found_count INTEGER NOT NULL DEFAULT 0,
  note TEXT,
  created_at TEXT NOT NULL,
  finished_at TEXT
);
CREATE TABLE IF NOT EXISTS places (
  job_id INTEGER NOT NULL REFERENCES jobs(id),
  place_key TEXT NOT NULL,
  name TEXT, category TEXT, address TEXT, country TEXT, phone TEXT, phone_digits TEXT, website TEXT, domain TEXT,
  rating REAL, reviews_count INTEGER, price_range TEXT, lat REAL, lng REAL, distance_km REAL, maps_url TEXT, plus_code TEXT, place_id TEXT,
  email TEXT, emails TEXT, instagram TEXT, facebook TEXT, linkedin TEXT, twitter_x TEXT, youtube TEXT, tiktok TEXT,
  whatsapp_number TEXT, whatsapp_source TEXT,
  enrich_status TEXT NOT NULL DEFAULT 'pending', enriched_at TEXT, scraped_at TEXT,
  summary TEXT, owner TEXT, team TEXT, research_status TEXT, researched_at TEXT, raw TEXT,
  PRIMARY KEY (job_id, place_key)
);
CREATE INDEX IF NOT EXISTS idx_places_phone ON places(phone_digits);
CREATE INDEX IF NOT EXISTS idx_places_domain ON places(domain);
CREATE INDEX IF NOT EXISTS idx_places_enrich ON places(enrich_status);
-- WhatsApp accounts used to verify numbers (one persistent browser profile each).
-- sent_today/day track the per-account daily cap; reset lazily when `day` rolls over.
CREATE TABLE IF NOT EXISTS wa_accounts (
  name TEXT PRIMARY KEY,
  added_at TEXT NOT NULL,
  last_used_at TEXT,
  sent_today INTEGER NOT NULL DEFAULT 0,
  day TEXT,
  disabled INTEGER NOT NULL DEFAULT 0
);
-- One row per phase per finished job: how long it took and how many units it handled.
-- The ETA (W4) is a rolling average over these, NOT a hardcoded constant, so the number
-- shown reflects this machine, this network and this delay setting. Kept forever — it is
-- a handful of bytes per job — and read newest-first with a LIMIT.
CREATE TABLE IF NOT EXISTS phase_rates (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  job_id INTEGER,
  phase TEXT NOT NULL,           -- scraping | enriching | researching | verifying_wa
  units INTEGER NOT NULL,        -- places scraped / sites crawled / numbers checked
  seconds REAL NOT NULL,         -- wall-clock the phase spent on those units
  recorded_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_phase_rates_phase ON phase_rates(phase, id DESC);
-- Per-job log history. `jobs.message` holds only the latest line (the progress bar reads
-- it and it is overwritten constantly); this is what actually happened, in order, so the
-- CRM's Logs dialog can show it without anyone opening data/agent.log on the PC.
-- `lane` is discovery | enrichment | whatsapp | job.
CREATE TABLE IF NOT EXISTS job_logs (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  job_id INTEGER NOT NULL,
  ts TEXT NOT NULL,
  lane TEXT NOT NULL DEFAULT 'job',
  level TEXT NOT NULL DEFAULT 'info',   -- info | warn | error
  message TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_job_logs_job ON job_logs(job_id, id);

-- Every link the Google Maps feed offered a job, and whether it was opened. Before this
-- the links lived only in memory, so a run that hit the Maps time limit with 37 of 237
-- unopened could only "search again" — now a `discovery_pending` re-run opens exactly
-- those 37 (CRM T172, 2026-08-25).
CREATE TABLE IF NOT EXISTS job_links (
  job_id INTEGER NOT NULL,
  key TEXT NOT NULL,
  href TEXT NOT NULL,
  name TEXT, rating REAL, reviews INTEGER, lat REAL, lng REAL,
  opened INTEGER NOT NULL DEFAULT 0,
  PRIMARY KEY (job_id, key)
);
"""


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def plus(number: str | None) -> str | None:
    """E.164 with the leading '+' (user directive 2026-08-23: every WhatsApp number shows
    one). Stored WITH the '+'; `wa.me/` links strip it again at render time because that
    URL form wants bare digits. Idempotent, and leaves anything non-numeric alone."""
    if not number:
        return None
    s = str(number).strip()
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    return f"+{digits}"


def fmt_team(team: Any) -> str:
    """team list[dict] → 'Name (Role) — phone / email; …' for CSV/Excel cells."""
    if not isinstance(team, list):
        return ""
    out = []
    for t in team:
        if not isinstance(t, dict):
            continue
        bits = t.get("name") or ""
        if t.get("role"):
            bits += f" ({t['role']})"
        contact = " / ".join(x for x in (t.get("phone"), t.get("email")) if x)
        if contact:
            bits += f" — {contact}"
        if bits:
            out.append(bits)
    return "; ".join(out)


class Store:
    EXPORT_COLS = [
        "name", "category", "phone", "whatsapp_number", "whatsapp_source", "wa_verified", "email", "emails", "website",
        "instagram", "facebook", "linkedin", "twitter_x", "youtube", "tiktok",
        "address", "country", "rating", "reviews_count", "price_range", "lat", "lng", "distance_km",
        "summary", "owner", "team", "maps_url", "place_id", "enrich_status", "job_id",
    ]

    def __init__(self, path: Path | None = None):
        self.path = path or settings.db_path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        # timeout + WAL: the web server's request handlers and the scrape worker thread each
        # open their own Store; WAL lets readers poll while the worker writes.
        self.conn = sqlite3.connect(self.path, timeout=30)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.executescript(SCHEMA)
        self._migrate()

    def _migrate(self) -> None:
        """Add columns introduced after a DB was first created (SQLite has no IF NOT EXISTS for ADD COLUMN)."""
        have = {r[1] for r in self.conn.execute("PRAGMA table_info(places)")}
        for col, typ in (("price_range", "TEXT"), ("country", "TEXT"), ("distance_km", "REAL"),
                         ("summary", "TEXT"), ("owner", "TEXT"), ("team", "TEXT"),
                         ("research_status", "TEXT"), ("researched_at", "TEXT"),
                         ("wa_verified", "TEXT"),        # 'yes' | 'no' | 'unknown'
                         ("wa_verified_at", "TEXT"), ("wa_verify_account", "TEXT"),
                         # WHY a crawl produced nothing: http_403 | http_<code> | dns |
                         # timeout | non_html | no_pages. Without this, enrich_status
                         # 'failed' is unexplainable — 8 of job #6's 9 failures were
                         # plain WAF 403s and looked identical to a broken site.
                         ("enrich_error", "TEXT"),
                         # HOW the home page was finally fetched: httpx | tls | browser.
                         # Shows which anti-bot tier rescued a lead — a 'tls'/'browser' value
                         # is a site that plain httpx could not read on its own.
                         ("enrich_via", "TEXT")):
            if col not in have:
                self.conn.execute(f"ALTER TABLE places ADD COLUMN {col} {typ}")
        # `changed_at` bumps on EVERY update to a place (enrichment verdict, socials, WA
        # result), so the agent can stream updates to the CRM mid-job instead of only new
        # rows — the CRM showed "64 / 248" while the agent had enriched 116 (job #14). The
        # WHEN guard stops the trigger re-firing on its own write.
        if "changed_at" not in have:
            self.conn.execute("ALTER TABLE places ADD COLUMN changed_at TEXT")
        self.conn.execute("""
            CREATE TRIGGER IF NOT EXISTS places_changed_at AFTER UPDATE ON places
            WHEN NEW.changed_at IS OLD.changed_at
            BEGIN
                UPDATE places SET changed_at = strftime('%Y-%m-%dT%H:%M:%f', 'now')
                WHERE rowid = NEW.rowid;
            END""")
        have = {r[1] for r in self.conn.execute("PRAGMA table_info(jobs)")}
        for col, typ in (
            ("phase", "TEXT"),            # queued | scraping | enriching | done | stopped | failed
            ("links_found", "INTEGER NOT NULL DEFAULT 0"),
            ("scraped_count", "INTEGER NOT NULL DEFAULT 0"),
            ("enrich_done", "INTEGER NOT NULL DEFAULT 0"),
            ("enrich_total", "INTEGER NOT NULL DEFAULT 0"),
            ("message", "TEXT"),
            ("stop_requested", "INTEGER NOT NULL DEFAULT 0"),
            ("do_enrich", "INTEGER NOT NULL DEFAULT 1"),
            ("headless", "INTEGER NOT NULL DEFAULT 1"),
            ("country", "TEXT"),
            ("reenrich_only", "INTEGER NOT NULL DEFAULT 0"),   # re-queued just to retry enrichment
            # Re-run that opens only the job_links never visited (Maps cap), no new search.
            ("discovery_pending", "INTEGER NOT NULL DEFAULT 0"),
            # JSON array of place_key: a re-run scoped to exactly the leads the user
            # had filtered in the CRM. NULL/absent = the whole job. Without this the
            # CRM's "Re-enrich (24)" silently re-ran everything.
            ("place_keys", "TEXT"),
            ("radius_km", "REAL"),           # optional: skip places farther than this from the centre
            ("center_lat", "REAL"), ("center_lng", "REAL"),   # resolved from Maps when the job starts
            ("window_start", "TEXT"), ("window_end", "TEXT"), # legacy clock window (API still honours it)
            ("max_minutes", "INTEGER"),      # stop scraping this many minutes after the job starts
            ("unique_new", "INTEGER NOT NULL DEFAULT 0"),      # skip places already scraped by ANY job
            ("skipped_known", "INTEGER NOT NULL DEFAULT 0"),
            ("do_research", "INTEGER NOT NULL DEFAULT 0"),     # LLM summary/owner/team pass
            ("research_done", "INTEGER NOT NULL DEFAULT 0"), ("research_total", "INTEGER NOT NULL DEFAULT 0"),
            ("do_wa_verify", "INTEGER NOT NULL DEFAULT 0"),    # verify each number against WhatsApp Web
            ("wa_verify_done", "INTEGER NOT NULL DEFAULT 0"), ("wa_verify_total", "INTEGER NOT NULL DEFAULT 0"),
            ("locations", "TEXT"),           # JSON [{location, radius_km, lat, lng}] — multi-area scrape
            ("skipped_far", "INTEGER NOT NULL DEFAULT 0"),
            ("started_at", "TEXT"),          # worker picked it up
            ("scrape_started_at", "TEXT"),
            ("enrich_started_at", "TEXT"),
            ("cloud_id", "INTEGER"),         # scrape_jobs.id when mirrored from the cloud queue
            ("cloud_kind", "TEXT"),          # 'saas' | 'crm' — which queue cloud_id belongs to
            # Phase deadlines, written by the worker when the job starts. Persisted (not
            # just held in the worker's memory) because the ETA is computed in the API
            # request handler and by the agent loop, both of which open their own Store.
            ("maps_deadline_at", "TEXT"),    # collect+scrape must stop by this instant
            ("enrich_deadline_at", "TEXT"),  # enrich/research/WA must stop by this instant
            ("research_started_at", "TEXT"),
            ("wa_started_at", "TEXT"),
            # ── per-lane outcome (2026-08-23) ───────────────────────────────────────
            # The three lanes now run concurrently, so one `phase` column cannot say how
            # each finished. Each lane writes ONLY its own three columns — that disjoint
            # ownership is what makes three threads sharing this DB safe, so keep it that
            # way when adding counters. `*_ok` is 1 only for a lane that genuinely ran out
            # of work; `*_reason` is a token (completed | maps_cap | stopped | wa_daily_cap
            # | wa_not_logged_in | no_targets | error:<detail>) the UI renders as prose.
            # Lane starts reuse scrape_started_at / enrich_started_at / wa_started_at.
            ("disc_ended_at", "TEXT"), ("disc_ok", "INTEGER"), ("disc_reason", "TEXT"),
            ("enr_ended_at", "TEXT"), ("enr_ok", "INTEGER"), ("enr_reason", "TEXT"),
            ("wa_ended_at", "TEXT"), ("wa_ok", "INTEGER"), ("wa_reason", "TEXT"),
            ("logs_synced_upto", "INTEGER NOT NULL DEFAULT 0"),
            # Leads in flight RIGHT NOW per lane (batch size while a batch runs, 0 between),
            # so the CRM can show done / in progress / queued instead of "1 / >= 1" (T170).
            ("enrich_active", "INTEGER NOT NULL DEFAULT 0"), ("wa_active", "INTEGER NOT NULL DEFAULT 0"),
        ):
            if col not in have:
                self.conn.execute(f"ALTER TABLE jobs ADD COLUMN {col} {typ}")
        # 'assumed_mobile' is retired (2026-08-23): a plain mobile number is a CANDIDATE for
        # verification, never a claim that the business is on WhatsApp. Existing rows become
        # 'unverified', which the UI deliberately renders with no tag at all.
        self.conn.execute(
            "UPDATE places SET whatsapp_source='unverified' WHERE whatsapp_source='assumed_mobile'")
        # jobs created by the CLI before `phase` existed: derive it from status so the UI can
        # label them; a stale 'running' (crashed run) becomes 'stopped'.
        self.conn.execute(
            "UPDATE jobs SET phase = CASE status WHEN 'running' THEN 'stopped' ELSE status END WHERE phase IS NULL"
        )
        self.conn.commit()

    def update_job(self, job_id: int, **fields: Any) -> None:
        if not fields:
            return
        cols = ",".join(f"{k}=?" for k in fields)
        self.conn.execute(f"UPDATE jobs SET {cols} WHERE id=?", [*fields.values(), job_id])
        self.conn.commit()

    # ── phase timing history (feeds the ETA; see eta.py) ─────────────────────
    def record_phase_rate(self, job_id: int | None, phase: str, units: int, seconds: float) -> None:
        """Remember that `phase` took `seconds` to handle `units` items.

        Called once per phase as it ends. Zero-unit or negative-duration runs are
        dropped: they carry no rate information and would poison the average
        (a phase that found nothing is not evidence that the work is instant).
        """
        if units <= 0 or seconds <= 0:
            return
        self.conn.execute(
            "INSERT INTO phase_rates(job_id, phase, units, seconds, recorded_at) VALUES (?,?,?,?,?)",
            (job_id, phase, int(units), float(seconds), now_iso()))
        self.conn.commit()

    def phase_rate(self, phase: str, window: int | None = None) -> float | None:
        """Rolling seconds-per-unit for `phase` over the last `window` recorded runs.

        Weighted by units (total seconds / total units) rather than averaging each
        run's rate, so a 3-place run cannot swing the estimate as hard as a 300-place
        one. Returns None when there is no history at all — the caller must then show
        "estimating…" instead of inventing a number.
        """
        window = window or settings.eta_history_jobs
        row = self.conn.execute(
            "SELECT SUM(units) AS u, SUM(seconds) AS s FROM "
            "(SELECT units, seconds FROM phase_rates WHERE phase=? ORDER BY id DESC LIMIT ?)",
            (phase, window)).fetchone()
        if not row or not row["u"]:
            return None
        return float(row["s"]) / float(row["u"])

    # ── job logs + per-lane bookkeeping (2026-08-23 lanes) ───────────────────────────
    def log(self, job_id: int | None, lane: str, message: str, level: str = "info") -> None:
        """Append one line to the job's log history. Never raises: a logging failure must
        not take down the lane it is describing."""
        if not job_id:
            return
        try:
            self.conn.execute(
                "INSERT INTO job_logs(job_id, ts, lane, level, message) VALUES (?,?,?,?,?)",
                (job_id, now_iso(), lane, level, str(message)[:1000]))
            self.conn.commit()
        except sqlite3.Error:                                    # noqa: BLE001
            log.debug("job_log write failed for job %s", job_id, exc_info=True)

    def logs(self, job_id: int, after_id: int = 0, limit: int = 500) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            "SELECT id, ts, lane, level, message FROM job_logs WHERE job_id=? AND id>? "
            "ORDER BY id LIMIT ?", (job_id, after_id, limit))]

    #: lane key -> (started column, ended column, ok column, reason column)
    LANE_COLS = {
        "discovery": ("scrape_started_at", "disc_ended_at", "disc_ok", "disc_reason"),
        "enrichment": ("enrich_started_at", "enr_ended_at", "enr_ok", "enr_reason"),
        "whatsapp": ("wa_started_at", "wa_ended_at", "wa_ok", "wa_reason"),
    }
    #: reasons that mean the lane finished its work rather than gave up
    OK_REASONS = ("completed", "no_targets")

    def lane_start(self, job_id: int, lane: str) -> None:
        started, ended, ok, reason = self.LANE_COLS[lane]
        # Clear a previous run's outcome so a resumed job does not show a stale reason.
        self.conn.execute(
            f"UPDATE jobs SET {started}=?, {ended}=NULL, {ok}=NULL, {reason}=NULL WHERE id=?",
            (now_iso(), job_id))
        self.conn.commit()

    def lane_disabled(self, job_id: int, lane: str) -> None:
        """Mark a lane this job will not run, and CLEAR its stamps.

        Clearing the start stamp is the point. On a re-run the row still carries the
        start time from the job's original pass, so a lane that is now switched off
        (Maps discovery during a re-enrich) looked "started but never ended" — i.e.
        running for ever, with a spinner, on a job that had finished.
        """
        started, ended, ok, reason_col = self.LANE_COLS[lane]
        self.conn.execute(
            f"UPDATE jobs SET {started}=NULL, {ended}=NULL, {ok}=NULL, {reason_col}='disabled' "
            f"WHERE id=?", (job_id,))
        self.conn.commit()

    def lane_end(self, job_id: int, lane: str, reason: str) -> None:
        started, ended, ok, reason_col = self.LANE_COLS[lane]
        self.conn.execute(
            f"UPDATE jobs SET {ended}=?, {ok}=?, {reason_col}=? WHERE id=?",
            (now_iso(), 1 if reason in self.OK_REASONS else 0, reason, job_id))
        self.conn.commit()

    def lanes(self, job_id: int) -> dict[str, dict[str, Any]]:
        """Per-lane {started_at, ended_at, ok, reason} for the info dialog."""
        r = self.conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
        if not r:
            return {}
        keys = r.keys()
        out = {}
        for lane, (s, e, ok, why) in self.LANE_COLS.items():
            out[lane] = {
                "started_at": r[s] if s in keys else None,
                "ended_at": r[e] if e in keys else None,
                "ok": None if (ok not in keys or r[ok] is None) else bool(r[ok]),
                "reason": r[why] if why in keys else None,
            }
        return out

    # ── lane input queues (the `places` table IS the queue) ──────────────────────────
    @staticmethod
    def _decode_json_cols(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """`emails`/`raw`/`team` are stored as JSON text. Consumers expect them decoded —
        `enrich.resolve_short_wa` does `row["raw"].get("maps_wa_links")` and dies with
        "'str' object has no attribute 'get'" on a raw row. Found by a live job, because
        the lane queues below originally returned rows straight from `SELECT *`."""
        for r in rows:
            for c in _JSON_COLS:
                try:
                    r[c] = json.loads(r[c]) if r[c] else ({} if c == "raw" else [])
                except (TypeError, json.JSONDecodeError):
                    r[c] = {} if c == "raw" else []
        return rows

    def _scope_clause(self, job_id: int) -> tuple[str, list[Any]]:
        """`AND place_key IN (...)` when the job is scoped to a subset, else nothing.

        A re-run started from the CRM carries the place_keys the user had FILTERED in
        the view. Ignoring it is not a cosmetic bug: "Re-enrich (24)" would re-crawl
        every lead in the job.
        """
        keys = self.job_place_keys(job_id)
        if not keys:
            return "", []
        return f" AND place_key IN ({','.join('?' * len(keys))})", list(keys)

    def pending_enrichment(self, job_id: int, limit: int = 10) -> list[dict[str, Any]]:
        """Leads discovery has written that enrichment has not taken yet."""
        scope, args = self._scope_clause(job_id)
        return self._decode_json_cols([dict(r) for r in self.conn.execute(
            "SELECT * FROM places WHERE job_id=? AND enrich_status='pending'" + scope
            + " ORDER BY rowid LIMIT ?", (job_id, *args, limit))])

    def pending_wa_verify(self, job_id: int, limit: int = 25) -> list[dict[str, Any]]:
        """Leads whose enrichment has RESOLVED (any outcome — a 403 site still has the
        Maps phone) that carry a number and have no verdict yet. 'yes'/'no' are final:
        re-checking them would burn the daily cap for nothing."""
        scope, args = self._scope_clause(job_id)
        return self._decode_json_cols([dict(r) for r in self.conn.execute(
            "SELECT * FROM places WHERE job_id=? AND enrich_status <> 'pending' "
            "AND (COALESCE(phone,'') <> '' OR COALESCE(whatsapp_number,'') <> '') "
            "AND COALESCE(wa_verified,'') NOT IN ('yes','no')" + scope
            + " ORDER BY rowid LIMIT ?", (job_id, *args, limit))])

    def job_place_keys(self, job_id: int) -> list[str] | None:
        """The subset this job is scoped to, or None for "every lead in the job"."""
        r = self.conn.execute("SELECT place_keys FROM jobs WHERE id=?", (job_id,)).fetchone()
        raw = r["place_keys"] if r and "place_keys" in r.keys() else None
        if not raw:
            return None
        try:
            keys = json.loads(raw)
        except (TypeError, json.JSONDecodeError):
            return None
        return [str(k) for k in keys] if isinstance(keys, list) and keys else None

    # ── job_links: what the Maps feed offered vs what was opened ───────────────
    def save_links(self, job_id: int, cards: list[Any]) -> None:
        """Upsert the feed cards of this run; `opened` is never reset here."""
        self.conn.executemany(
            "INSERT INTO job_links(job_id, key, href, name, rating, reviews, lat, lng) "
            "VALUES (?,?,?,?,?,?,?,?) ON CONFLICT(job_id, key) DO UPDATE SET href=excluded.href, "
            "name=COALESCE(excluded.name, job_links.name)",
            [(job_id, c.key, c.href, c.name, c.rating, c.reviews_count, c.lat, c.lng) for c in cards])
        self.conn.commit()

    def mark_link_opened(self, job_id: int, key: str) -> None:
        self.conn.execute("UPDATE job_links SET opened=1 WHERE job_id=? AND key=?", (job_id, key))
        self.conn.commit()

    def pending_links(self, job_id: int) -> list[dict[str, Any]]:
        """Feed links this job never opened, in feed order."""
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM job_links WHERE job_id=? AND opened=0 ORDER BY rowid", (job_id,))]

    def link_counts(self, job_id: int) -> tuple[int, int]:
        """(offered, opened) for the discovery stats."""
        r = self.conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(opened),0) FROM job_links WHERE job_id=?", (job_id,)).fetchone()
        return int(r[0] or 0), int(r[1] or 0)

    def count_places(self, job_id: int) -> int:
        r = self.conn.execute("SELECT COUNT(*) FROM places WHERE job_id=?", (job_id,)).fetchone()
        return int(r[0] or 0)

    def count_pending_enrichment(self, job_id: int) -> int:
        """How many leads still await enrichment — respecting the job's place_keys scope, so
        a scoped re-enrich counts only its subset. This is what the enrichment lane's `total`
        should measure against: on a 21-lead re-enrich the bar must read "/21", not "/180".
        Leads with no website are not enrichable and are left out (T163) — otherwise the CRM
        card's done + outstanding never adds up to its total."""
        scope, args = self._scope_clause(job_id)
        r = self.conn.execute(
            "SELECT COUNT(*) FROM places WHERE job_id=? AND enrich_status='pending'"
            " AND website IS NOT NULL AND website<>''" + scope,
            (job_id, *args)).fetchone()
        return int(r[0] or 0)

    def count_enriched(self, job_id: int) -> int:
        """Enrichable leads already past 'pending' in this job's scope — the enrichment
        lane's starting `done`. A lane that resumes after an agent restart (orphan re-queue)
        used to start its numerator at 0 while the results kept every lead the interrupted
        run had enriched: job #14 (2026-08-25) read "1 / ≥ 1" beside 45 emails found."""
        scope, args = self._scope_clause(job_id)
        r = self.conn.execute(
            "SELECT COUNT(*) FROM places WHERE job_id=? AND enrich_status<>'pending'"
            " AND website IS NOT NULL AND website<>''" + scope,
            (job_id, *args)).fetchone()
        return int(r[0] or 0)

    def count_wa_done(self, job_id: int) -> int:
        """Numbers with a final WhatsApp verdict (yes/no) in scope."""
        scope, args = self._scope_clause(job_id)
        r = self.conn.execute(
            "SELECT COUNT(*) FROM places WHERE job_id=? AND wa_verified IN ('yes','no')" + scope,
            (job_id, *args)).fetchone()
        return int(r[0] or 0)

    def count_wa_pending(self, job_id: int) -> int:
        """Same predicate as pending_wa_verify(), as a count: the WhatsApp lane's pipeline."""
        scope, args = self._scope_clause(job_id)
        r = self.conn.execute(
            "SELECT COUNT(*) FROM places WHERE job_id=? AND enrich_status <> 'pending' "
            "AND (COALESCE(phone,'') <> '' OR COALESCE(whatsapp_number,'') <> '') "
            "AND COALESCE(wa_verified,'') NOT IN ('yes','no')" + scope,
            (job_id, *args)).fetchone()
        return int(r[0] or 0)

    def stop_requested(self, job_id: int) -> bool:
        r = self.conn.execute("SELECT stop_requested FROM jobs WHERE id=?", (job_id,)).fetchone()
        return bool(r and r[0])

    def next_queued_job(self) -> sqlite3.Row | None:
        return self.conn.execute(
            "SELECT * FROM jobs WHERE phase='queued' ORDER BY id LIMIT 1"
        ).fetchone()

    def queued_jobs(self) -> list[sqlite3.Row]:
        return self.conn.execute("SELECT * FROM jobs WHERE phase IN ('queued','waiting') ORDER BY id").fetchall()

    # ── WhatsApp verification: numbers + account rotation with a per-account daily cap ──
    def set_wa_verify(self, job_id: int, place_key: str, status: str, account: str | None,
                      wa_number: str | None = None, prior_source: str | None = None) -> None:
        self.conn.execute(
            "UPDATE places SET wa_verified=?, wa_verified_at=?, wa_verify_account=? WHERE job_id=? AND place_key=?",
            (status, now_iso(), account, job_id, place_key))
        # On a confirmed hit, promote the verified number into whatsapp_number and mark the
        # source 'verified' (replacing an 'unverified' candidate). On a miss, drop the
        # candidate so an unverified number never lingers as if it were a WhatsApp.
        if status == "yes" and wa_number:
            self.conn.execute(
                "UPDATE places SET whatsapp_number=?, whatsapp_source='verified' WHERE job_id=? AND place_key=?",
                (plus(wa_number), job_id, place_key))
        elif status == "no" and prior_source in ("unverified", "assumed_mobile"):
            self.conn.execute(
                "UPDATE places SET whatsapp_number=NULL, whatsapp_source=NULL WHERE job_id=? AND place_key=?",
                (job_id, place_key))
        self.conn.commit()

    def list_wa_accounts(self) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute("SELECT * FROM wa_accounts ORDER BY name")]

    def add_wa_account(self, name: str) -> None:
        # Upsert + re-enable: a fresh wa-login clears a prior 'disabled' flag (e.g. one
        # set when a headless verify misread the session as logged-out).
        self.conn.execute(
            "INSERT INTO wa_accounts(name, added_at) VALUES (?, ?) "
            "ON CONFLICT(name) DO UPDATE SET disabled=0", (name, now_iso()))
        self.conn.commit()

    def remove_wa_account(self, name: str) -> None:
        self.conn.execute("DELETE FROM wa_accounts WHERE name=?", (name,))
        self.conn.commit()

    def _roll_day(self, today: str) -> None:
        """Zero every account's counter whose stored day is not today (lazy daily reset)."""
        self.conn.execute("UPDATE wa_accounts SET sent_today=0, day=? WHERE day IS NULL OR day<>?",
                          (today, today))
        self.conn.commit()

    def wa_capacity(self, cap: int, today: str) -> int:
        """Total remaining checks across all enabled accounts for `today`."""
        self._roll_day(today)
        rows = self.conn.execute("SELECT sent_today FROM wa_accounts WHERE disabled=0").fetchall()
        return sum(max(0, cap - int(r[0])) for r in rows)

    def pick_wa_account(self, cap: int, today: str) -> str | None:
        """Enabled account with remaining cap today, least-recently-used first (rotation)."""
        self._roll_day(today)
        row = self.conn.execute(
            "SELECT name FROM wa_accounts WHERE disabled=0 AND sent_today<? "
            "ORDER BY last_used_at IS NULL DESC, last_used_at ASC LIMIT 1", (cap,)).fetchone()
        return row[0] if row else None

    def bump_wa_account(self, name: str, today: str) -> None:
        self.conn.execute(
            "UPDATE wa_accounts SET sent_today=sent_today+1, last_used_at=?, day=? WHERE name=?",
            (now_iso(), today, name))
        self.conn.commit()

    def export_xlsx(self, job_id: int | None, out: Path | None = None, unique: bool = False) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        rows = self.export_rows(job_id, unique)
        settings.export_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        out = out or settings.export_dir / (f"leads-job{job_id}-{stamp}.xlsx" if job_id else f"leads-all-{stamp}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(self.EXPORT_COLS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for r in rows:
            r = dict(r)
            r["emails"] = "; ".join(r.get("emails") or []); r["team"] = fmt_team(r.get("team"))
            ws.append([r.get(c) for c in self.EXPORT_COLS])
        widths = {"name": 34, "address": 50, "website": 32, "email": 30, "emails": 36, "maps_url": 18,
                  "instagram": 30, "facebook": 30, "linkedin": 30, "twitter_x": 26, "youtube": 26}
        for i, c in enumerate(self.EXPORT_COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(out)
        return out

    # ── jobs ────────────────────────────────────────────────────────────────
    def create_job(self, query: str, location: str | None, max_places: int, delay_sec: float,
                   phase: str = "scraping", do_enrich: bool = True, headless: bool = True,
                   country: str | None = None, radius_km: float | None = None,
                   window_start: str | None = None, window_end: str | None = None,
                   center_lat: float | None = None, center_lng: float | None = None,
                   max_minutes: int | None = None, unique_new: bool = False) -> int:
        cur = self.conn.execute(
            "INSERT INTO jobs(query, location, max_places, delay_sec, status, created_at, phase, do_enrich, headless, "
            "country, radius_km, window_start, window_end, center_lat, center_lng, max_minutes, unique_new) "
            "VALUES (?,?,?,?,'running',?,?,?,?,?,?,?,?,?,?,?,?)",
            (query, location, max_places, delay_sec, now_iso(), phase, int(do_enrich), int(headless), country,
             radius_km, window_start, window_end, center_lat, center_lng, max_minutes, int(unique_new)),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def finish_job(self, job_id: int, status: str = "done", note: str | None = None) -> None:
        n = self.conn.execute("SELECT COUNT(*) FROM places WHERE job_id=?", (job_id,)).fetchone()[0]
        self.conn.execute(
            "UPDATE jobs SET status=?, note=?, found_count=?, finished_at=? WHERE id=?",
            (status, note, n, now_iso(), job_id),
        )
        self.conn.commit()

    def get_job(self, job_id: int) -> sqlite3.Row | None:
        return self.conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()

    def list_jobs(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT j.*, (SELECT COUNT(*) FROM places p WHERE p.job_id=j.id) AS places FROM jobs j ORDER BY id DESC"
        ).fetchall()

    # ── places ──────────────────────────────────────────────────────────────
    def known_place_keys(self, job_id: int) -> set[str]:
        return {r[0] for r in self.conn.execute("SELECT place_key FROM places WHERE job_id=?", (job_id,))}

    def all_place_keys(self) -> set[str]:
        """Every place ever scraped, across all jobs — used by 'only new businesses' jobs."""
        return {r[0] for r in self.conn.execute("SELECT DISTINCT place_key FROM places")}

    def leads_all(self, unique: bool = True) -> list[dict[str, Any]]:
        """All scraped leads across jobs, newest first, joined with their job's query/location.
        `unique` keeps the most recently scraped row per place."""
        q = ("SELECT p.*, j.query AS job_query, j.location AS job_location FROM places p "
             "JOIN jobs j ON j.id = p.job_id ORDER BY p.scraped_at DESC, p.rowid DESC")
        rows = [dict(r) for r in self.conn.execute(q)]
        for r in rows:
            for c in _JSON_COLS:
                try:
                    r[c] = json.loads(r[c]) if r[c] else ({} if c == "raw" else [])
                except (TypeError, json.JSONDecodeError):
                    r[c] = {} if c == "raw" else []
        if not unique:
            return rows
        seen: set[str] = set()
        out = []
        for r in rows:            # newest first → first occurrence wins
            if r["place_key"] in seen:
                continue
            seen.add(r["place_key"])
            out.append(r)
        return out

    def upsert_place(self, p: Place) -> None:
        row = p.to_row()
        vals = [json.dumps(row[c], ensure_ascii=False) if c in _JSON_COLS else row[c] for c in PLACE_COLS]
        placeholders = ",".join("?" for _ in PLACE_COLS)
        updates = ",".join(f"{c}=excluded.{c}" for c in PLACE_COLS if c not in ("job_id", "place_key"))
        self.conn.execute(
            f"INSERT INTO places({','.join(PLACE_COLS)}) VALUES ({placeholders}) "
            f"ON CONFLICT(job_id, place_key) DO UPDATE SET {updates}",
            vals,
        )
        self.conn.commit()

    def update_enrichment(self, job_id: int, place_key: str, fields: dict[str, Any]) -> None:
        fields = dict(fields)
        if "emails" in fields and not isinstance(fields["emails"], str):
            fields["emails"] = json.dumps(fields["emails"], ensure_ascii=False)
        cols = ",".join(f"{k}=?" for k in fields)
        self.conn.execute(
            f"UPDATE places SET {cols} WHERE job_id=? AND place_key=?",
            [*fields.values(), job_id, place_key],
        )
        self.conn.commit()

    def places_changed_since(self, job_id: int, since: str, upto_rowid: int) -> tuple[list[dict[str, Any]], str]:
        """Places already streamed (rowid <= upto_rowid) that changed after `since`, plus the
        new watermark. New rows go through places_after(); this catches the enrichment and
        WhatsApp verdicts written onto rows the CRM already holds."""
        rows = self.conn.execute(
            "SELECT * FROM places WHERE job_id=? AND rowid<=? AND changed_at IS NOT NULL "
            "AND changed_at>? ORDER BY changed_at", (job_id, upto_rowid, since)).fetchall()
        out = self._decode_json_cols([dict(r) for r in rows])
        top = max((str(r["changed_at"]) for r in rows), default=since)
        return out, top

    def places_after(self, job_id: int, after_rowid: int) -> tuple[list[dict[str, Any]], int]:
        """Places saved since `after_rowid`, plus the new watermark.

        Lets the agent stream leads to the CRM while a job is still running, so a job
        that is stopped or hits its time limit has already delivered what it found
        instead of surfacing zero. `places` has no surrogate key, so rowid is the
        watermark; rows only ever get appended during scraping.
        """
        rows = self.conn.execute(
            "SELECT rowid AS rid, * FROM places WHERE job_id=? AND rowid>? ORDER BY rowid",
            (job_id, after_rowid)).fetchall()
        out: list[dict[str, Any]] = []
        top = after_rowid
        for r in rows:
            d = dict(r)
            top = max(top, int(d.pop("rid")))
            out.append(d)
        return out, top

    def places(self, job_id: int | None = None, enrich_status: str | None = None) -> list[dict[str, Any]]:
        q = "SELECT * FROM places WHERE 1=1"
        args: list[Any] = []
        if job_id is not None:
            q += " AND job_id=?"
            args.append(job_id)
        if enrich_status is not None:
            q += " AND enrich_status=?"
            args.append(enrich_status)
        q += " ORDER BY job_id, rowid"
        rows = [dict(r) for r in self.conn.execute(q, args)]
        for r in rows:
            for c in _JSON_COLS:
                try:
                    r[c] = json.loads(r[c]) if r[c] else ({} if c == "raw" else [])
                except (TypeError, json.JSONDecodeError):
                    r[c] = {} if c == "raw" else []
        return rows

    def stats(self) -> dict[str, Any]:
        c = self.conn
        total = c.execute("SELECT COUNT(*) FROM places").fetchone()[0]

        def cnt(where: str) -> int:
            return c.execute(f"SELECT COUNT(*) FROM places WHERE {where}").fetchone()[0]

        return {
            "jobs": c.execute("SELECT COUNT(*) FROM jobs").fetchone()[0],
            "places": total,
            "with_phone": cnt("phone IS NOT NULL AND phone <> ''"),
            "with_website": cnt("website IS NOT NULL AND website <> ''"),
            "with_email": cnt("email IS NOT NULL AND email <> ''"),
            "with_instagram": cnt("instagram IS NOT NULL"),
            "with_facebook": cnt("facebook IS NOT NULL"),
            "with_linkedin": cnt("linkedin IS NOT NULL"),
            "with_twitter_x": cnt("twitter_x IS NOT NULL"),
            "wa_maps_link": cnt("whatsapp_source='maps_link'"),
            "wa_link": cnt("whatsapp_source='wa_link'"),
            # 'assumed_mobile' was retired 2026-08-23 — an unverified number is a
            # CANDIDATE for verification, never a claim that the business is on WhatsApp.
            "wa_unverified": cnt("whatsapp_source='unverified'"),
            "wa_verified": cnt("whatsapp_source='verified'"),
            "enrich_pending": cnt("enrich_status='pending'"),
            "enrich_done": cnt("enrich_status='done'"),
            "enrich_thin": cnt("enrich_status='thin'"),
            "enrich_failed": cnt("enrich_status='failed'"),
            "enrich_no_website": cnt("enrich_status='no_website'"),
            "unique_phones": c.execute("SELECT COUNT(DISTINCT phone_digits) FROM places WHERE phone_digits IS NOT NULL").fetchone()[0],
        }

    # ── export ──────────────────────────────────────────────────────────────
    def export_rows(self, job_id: int | None, unique: bool = False) -> list[dict[str, Any]]:
        """Rows for export. `unique` keeps the first place per website domain — collapses
        chains (one business, many branches) into a single lead; rows without a website stay."""
        rows = self.places(job_id)
        if not unique:
            return rows
        seen: set[str] = set()
        out = []
        for r in rows:
            d = r.get("domain")
            if d:
                if d in seen:
                    continue
                seen.add(d)
            out.append(r)
        return out

    def export_leads(self, fmt: str, unique: bool = True) -> Path:
        """Export the cross-job leads list (place-level unique)."""
        rows = self.leads_all(unique)
        cols = self.EXPORT_COLS + ["scraped_at", "job_query", "job_location"]
        settings.export_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        out = settings.export_dir / f"all-leads-{stamp}.{fmt}"
        if fmt == "json":
            out.write_text(json.dumps([{c: r.get(c) for c in cols} for r in rows],
                                      ensure_ascii=False, indent=2), encoding="utf-8")
        elif fmt == "csv":
            with out.open("w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
                w.writeheader()
                for r in rows:
                    r = dict(r)
                    r["emails"] = "; ".join(r.get("emails") or []); r["team"] = fmt_team(r.get("team"))
                    w.writerow({c: r.get(c) for c in cols})
        else:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "All leads"
            ws.append(cols)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
            for r in rows:
                r = dict(r)
                r["emails"] = "; ".join(r.get("emails") or []); r["team"] = fmt_team(r.get("team"))
                ws.append([r.get(c) for c in cols])
            for i, c in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(i)].width = {"name": 34, "address": 50, "website": 32,
                    "email": 30, "emails": 36, "instagram": 30, "facebook": 30, "linkedin": 30,
                    "scraped_at": 20, "job_query": 20, "job_location": 20}.get(c, 16)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            wb.save(out)
        return out

    def export(self, job_id: int | None, fmt: str, out: Path | None = None, unique: bool = False) -> Path:
        rows = self.export_rows(job_id, unique)
        settings.export_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        name = f"leads-job{job_id}-{stamp}" if job_id else f"leads-all-{stamp}"
        out = out or settings.export_dir / f"{name}.{fmt}"
        if fmt == "json":
            out.write_text(json.dumps([{c: r.get(c) for c in self.EXPORT_COLS} for r in rows],
                                      ensure_ascii=False, indent=2), encoding="utf-8")
        else:
            with out.open("w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=self.EXPORT_COLS, extrasaction="ignore")
                w.writeheader()
                for r in rows:
                    r = dict(r)
                    r["emails"] = "; ".join(r.get("emails") or []); r["team"] = fmt_team(r.get("team"))
                    w.writerow({c: r.get(c) for c in self.EXPORT_COLS})
        return out

    def close(self) -> None:
        self.conn.close()
